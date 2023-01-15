import openpyxl
from openpyxl.styles import Font

async def save_db_to_excel(chat, accounts):
    wb = openpyxl.Workbook()
    
    sheet = wb.active

    sheet['A1'] = 'Название группы'
    sheet['A2'] = chat['name']
    sheet['A1'].font = Font(bold=True, size=14)

    sheet['B1'] = 'Ссылка на группу'
    sheet['B2'] = chat['link']
    sheet['B1'].font = Font(bold=True, size=14)

    sheet['C1'] = 'Пользователи'
    sheet['C1'].font = Font(bold=True, size=14)

    sheet['D1'] = 'Статус'
    sheet['D1'].font = Font(bold=True, size=14)

    

    for i in range(len(chat['users'])):
        sheet.cell(row = i + 2, column= 3).value = chat['users'][i]
        # sheet.cell(row = i + 2, column= 4).value = accounts[0].get_entity(chat['users'][i])
        ent = await accounts[0].get_entity(chat['users'][i])
        print(ent)

    wb.save("base.xlsx")